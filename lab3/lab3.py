# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import Cell
import pdfkit
from xlsx2html import xlsx2html
import io

config = pdfkit.configuration(wkhtmltopdf="C:\\Users\\Дзаша\\wkhtmltox\\bin\\wkhtmltopdf.exe")
wb = openpyxl.load_workbook(filename = 'blank.xlsx')
sheet=wb['TDSheet']
sheet['B2'].value="АО Стоун Банк"
sheet['E5'].value="7722737363"
sheet['O5'].value="123456789"
sheet['AD2'].value="9876556"
sheet['AD3'].value="71301017200000000600"
sheet['B6'].value="ОА ВАСИЛЕК"
sheet['AD5'].value="45837483759847389729"
sheet['G14'].value="-ОА ВАСИЛЕК, ИИН 432352534, КПП 325234545, 3523543, г.Москва, ул.Гагарина, д.5, тел.: 89124834748"
sheet['G17'].value="-Попов А.М., тел.89128724453"
sheet['B20'].value="№3847783 от 20.05.2019"

print('Количество услуг-')
count=int(input())
buf=[]
for i in range(count+1):
    if i>=1:
        buf+=str(i)
        buf+='<br>'
    elif i==count:
        buf+=str(i)        
number=''.join(buf)

sheet['B23'].value=number

buf=[]
name=''
for i in range(count):
    print('Название услуги ',i+1)
    name=input()
    buf+=name
    buf+='<br>'
name=''.join(buf)
sheet['D23'].value=name

buf1=[]
inp=''
buf2=[]
total=0
summ=0
cnt=0
buf3=[]
for i in range(count):
    print('Количество ',i+1,'услуги')
    inp=input()
    cnt=float(inp)
    buf1+=inp
    buf1+='<br>'
    print('Цена',' ',i+1,'услуги')
    inp=input()
    summ=cnt*float(inp)
    buf2+=inp
    buf2+='<br>'
    buf3+=str(summ)
    buf3+='<br>'
    total+=summ
count2=''.join(buf1)
sheet['Z23'].value=count2
price=''.join(buf2)
sheet['AF23'].value=price
amount=''.join(buf3)
sheet['AK23'].value=amount

print("Счет на оплату № ...  от ...20__ г. -")
inp=input()
sheet['B10'].value=inp

sheet['AL25'].value=total

val=sheet['AK26'].value
print(val,end='')
sheet.merge_cells('AD26:AK26')
sheet['AD26'].value=val
Nds=float(input())
sheet['AL26'].value=Nds

val=sheet['AK27'].value
sheet.merge_cells('AD27:AK27')
sheet['AD27'].value=val
sheet['AL27'].value=Nds+total

total+=Nds

buf=[]
buf+="Всего наименований, на сумму "
buf+=str(total)
buf+="руб."
inp=''.join(buf)
sheet['B28'].value=inp

print("Итого=",total)
print("словами")
inp=input()
sheet['B29'].value=inp

sheet['M37'].value="Петров А.А."
sheet['AJ37'].value="Иванов Б.Б."

wb.save('blank2.xlsx')

xlsx2html('blank2.xlsx', 'bill.html')
with io.open('bill.html', 'r', encoding='ANSI') as f:
    text = f.read()
with io.open('bill.html', 'w', encoding='utf8') as f:
    f.write(text)
pdfkit.from_file('bill.html','total.pdf', configuration=config)
